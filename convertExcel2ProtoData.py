# -*- coding: utf-8 -*-

#import sys
#reload(sys)
#sys.setdefaultencoding("utf-8")

from openpyxl import load_workbook
import datetime
import time
import types
import os
import importlib

# google protobuf 生成文件路径
PYTHON_PROTO_PATH = "pb"

# excel文件后缀
EXCEL_FILE_SUFFIX = ".xlsx"

class Excel2ProtoBuf:
    """用于解析，保存Excel中数据
    """
    excelFile = None
    head = []
    filedProcessors = {}
    itemTable = None

    def __init__(self):
        pass

    def parseExcel(self, excelFile, module):
        """
        """
        self.excelFile = excelFile
        self.itemTable = module()
        wb = load_workbook(filename = self.excelFile, read_only=True)
        ws = wb.get_active_sheet()

        isHead = True
        for row in ws.iter_rows():
            if isHead:
                isHead = False
                for cell in row:
                    self.head.append(cell.value)
                continue
            
            cellItem = self.itemTable.item.add()
            for fieldName, cell in zip(self.head, row):
                try:
                    data = self._processCell(fieldName, cell.value)
                    setattr(cellItem, fieldName, data)
                except Exception as e:
                    print(u"字段%s表中值为：%s 原因：%s" % (fieldName, cell.value, e.message()))

    def _processCell(self, fieldName, data):
        """process cell data"""
        if isinstance(data, datetime.datetime):
            return self._datetimeProcessor(data)

        processor = self.filedProcessors[fieldName] if fieldName in self.filedProcessors else None
        if processor is not None:
            return processor(data)

        return data

    def _datetimeProcessor(self, data):
        """convert datetime.datetime into unix timestamp"""
        assert isinstance(data, datetime.datetime)
        return int(time.mktime(data.timetuple()))

    def registerFieldProcessor(self, specialField, processor):
        """注册特殊字段处理函数"""
        assert(isinstance(processor, types.FunctionType))
        if specialField in self.filedProcessors:
            print(u"字段%s已经注册了处理函数，将覆盖原函数!")

        self.filedProcessors[specialField] = processor

    def saveToFile(self, saveFile):
        """将数据保存到文件
        """
        with open(saveFile, "wb") as io:
            io.write(self.itemTable.SerializeToString())
        print("%s被转换保存为%s!" % (self.excelFile, saveFile))

    def parseFromFile(self, fileName):
        """parse data construct protobuf struct"""
        self.excelFile = fileName
        with open(fileName, "rb") as io:
            self.itemTable.ParseFromString(io.read())

        for it in self.itemTable.item:
            print(it)

def exportAll(excelPath, savePath):
    """将指定目录下Excel文件转换为protobuf二进制数据"""
    
    protoModule = None
    itemTable = Excel2ProtoBuf()

    for excelFile in os.listdir(excelPath):
        if not excelFile.endswith(EXCEL_FILE_SUFFIX):
            continue

        print(excelFile)
        try:
            protoModuleName = excelFile.rstrip(EXCEL_FILE_SUFFIX) + "_pb2"
            protoModule = importlib.import_module("pb." + protoModuleName)
            protoTable = excelFile.rstrip(EXCEL_FILE_SUFFIX) + "Table"
            protoModule = getattr(protoModule, protoTable)
        except Exception as e:
            print("导入google protobuf模块失败！name: [%s]\n原因:%s\n" %
                (excelFile, e.message()))
            continue

        itemTable.parseExcel(excelPath + "/" + excelFile, protoModule)
        itemTable.saveToFile(savePath + "/" + excelFile + ".pb")

def main():
    exportAll("excel", "data")

if __name__ == '__main__':
    main()
