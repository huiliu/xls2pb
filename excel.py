# -*- encoding: utf-8 -*-
from openpyxl import load_workbook
import item_pb2
import datetime
import time
import types

def main():
    itemTable = Excel2ProtoBuf()
    itemTable.parseExcel("item.xlsx")
    itemTable.saveToFile("item.pb")
    itemTable.parseFromFile("item.pb")

class Excel2ProtoBuf:
    """用于解析，保存Excel中数据
    """
    excelFile = None
    head = []
    filedProcessors = {}
    itemTable = item_pb2.Items()

    def __init__(self):
        pass

    def parseExcel(self, excelFile):
        self.excelFile = excelFile
        wb = load_workbook(filename = excelFile, read_only=True)
        ws = wb.get_active_sheet()

        isHead = True
        for row in ws.iter_rows():
            if isHead:
                isHead = False
                for cell in row:
                    self.head.append(cell.value)
                continue
            
            cellItem = self.itemTable.item.add()
            for fieldName, cell in zip(self.head, row):
                try:
                    data = self._processCell(fieldName, cell.value)
                    setattr(cellItem, fieldName, data)
                except Exception as e:
                    print(u"字段%s表中值为：%s 原因：%s" % (fieldName, cell.value, e.message))

    def _processCell(self, fieldName, data):
        """process cell data"""
        if isinstance(data, datetime.datetime):
            return self._datetimeProcessor(data)

        processor = self.filedProcessors[fieldName] if fieldName in self.filedProcessors else None
        if processor is not None:
            return processor(data)

        return data

    def _datetimeProcessor(self, data):
        """convert datetime.datetime into unix timestamp"""
        assert isinstance(data, datetime.datetime)
        return int(time.mktime(data.timetuple()))

    def registerFieldProcessor(self, specialField, processor):
        """注册特殊字段处理函数"""
        assert(isinstance(processor, types.FunctionType))
        if specialField in self.filedProcessors:
            print(u"字段%s已经注册了处理函数，将覆盖原函数!")

        self.filedProcessors[specialField] = processor

    def saveToFile(self, saveFile):
        with open(saveFile, "wb") as io:
            io.write(self.itemTable.SerializeToString())

    def parseFromFile(self, fileName):
        """parse data construct protobuf struct"""
        self.excelFile = fileName
        with open(fileName, "rb") as io:
            self.itemTable.ParseFromString(io.read())

        for it in self.itemTable.item:
            print(it)

if __name__ == '__main__':
    main()
